from __future__ import annotations

from pathlib import Path
from typing import Sequence

import openpyxl

from src.parsers.base import DocumentParser, ParsedDocument


class ExcelParser(DocumentParser):
    def __init__(
        self,
        sheet_names: Sequence[str] | None = None,
        max_rows: int = 10000,
    ) -> None:
        self.sheet_names = sheet_names
        self.max_rows = max_rows

    def parse(self, path: Path) -> ParsedDocument:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        try:
            sheets_to_parse = self.sheet_names if self.sheet_names else wb.sheetnames
            sheets_text = []
            parsed_names = []
            for sheet in sheets_to_parse:
                if sheet not in wb.sheetnames:
                    continue
                ws = wb[sheet]
                all_rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= self.max_rows:
                        break
                    cells = [str(c).replace("\n", " ").strip() if c is not None else "" for c in row]
                    # Skip fully empty rows
                    if not any(cells):
                        continue
                    all_rows.append(cells)
                if all_rows:
                    # Filter out empty columns — find columns that have at least one non-empty value
                    num_cols = max(len(r) for r in all_rows)
                    non_empty_cols = set()
                    for row in all_rows:
                        for i, cell in enumerate(row):
                            if cell.strip():
                                non_empty_cols.add(i)
                    # Keep only columns with data, in order
                    col_indices = sorted(non_empty_cols) if non_empty_cols else list(range(min(num_cols, 5)))
                    filtered = [[row[i] if i < len(row) else "" for i in col_indices] for row in all_rows]
                    # Build markdown table
                    header = filtered[0]
                    num_cols_filtered = len(header)
                    lines = [
                        "| " + " | ".join(header) + " |",
                        "| " + " | ".join("---" for _ in range(num_cols_filtered)) + " |",
                    ]
                    for row in filtered[1:]:
                        lines.append("| " + " | ".join(row) + " |")
                    sheets_text.append(f"## {sheet}\n" + "\n".join(lines))
                    parsed_names.append(sheet)
            return ParsedDocument(
                content="\n\n".join(sheets_text),
                metadata={"sheets": len(parsed_names), "sheet_names": parsed_names},
                source_path=str(path),
                file_type="excel",
            )
        finally:
            wb.close()
